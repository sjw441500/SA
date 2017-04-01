#-*- coding: utf-8 -*-
import xlrd
import xlwt
from xlutils.copy import copy
#import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date, datetime
import matplotlib.pyplot as plt
import re
color=['#FFFF00','#FFDEAD','#FF6A6A','#FF34B3','#EEC591','#DB7093','#BFEFFF','#AB82FF','#98FB98','#36648B']


#codes below written by zyy
def get_mediatoid(fname=r"./xls/id.xlsx", sheetindex=1, mediacol=2, idcol=3):
    '''

    :param fname: the name of the file containing a sheet that shows that relationship between media and id
    :param sheetindex:
    :param mediacol:
    :param idcol:
    :return: a dict with key=media and value=id(str)
    '''
    bk = xlrd.open_workbook(fname)
    mediatoid = bk.sheet_by_index(sheetindex)

    media_id = {}
    for row in xrange(mediatoid.nrows):
        if mediatoid.cell(row, mediacol).ctype == 1 and mediatoid.cell(row, idcol).ctype == 2:
            media = mediatoid.cell(row, mediacol).value
            id = str(int(mediatoid.cell(row, idcol).value))
            media_id[media] = id
    return media_id


def get_urltomedia(fname=r"./xls/id.xlsx", sheetindex=0, mediacol=1, urlcol=2):
    '''

    :param fname: the name of the file containing a sheet that shows that relationship between url and media
    :param sheetindex:
    :param mediacol:
    :param urlcol:
    :return:a dict with key=url keyword and value=media
    '''
    bk = xlrd.open_workbook(fname)
    urltomedia = bk.sheet_by_index(sheetindex)

    url_media = {}
    for row in xrange(urltomedia.nrows):
        if urltomedia.cell(row, mediacol).ctype == 1 and urltomedia.cell(row, urlcol).ctype == 1:
            media = urltomedia.cell(row, mediacol).value
            url = urltomedia.cell(row, urlcol).value
            url_media[url] = media
    return url_media


def get_urltoid():
    '''

    :return: a dict with key=url keyword and value=id(str)
    '''
    url_media = get_urltomedia()
    media_id = get_mediatoid()
    urlKeys = url_media.keys()

    url_id = {}
    for urlkey in urlKeys:
        media = url_media[urlkey]
        id = media_id.get(media)
        url_id[urlkey] = id

    return url_id


def m_idtoarea(fname, sheetindex, m_idcol, areacol):
    '''

    :param fname: a excel file name.default r"id.xlsx"
    :param sheetindex: the index of the sheet that has the relationship between area and media_id.default 3
    :param m_idcol:the column of media_id.default 3
    :param areacol:the column of area.default 0.  !!area must be expressed by a specific int
    :return:a dict with key=media-id and value=area(int)
    '''
    if fname is None:
        fname = r"./xls/id.xlsx"
    wb = xlrd.open_workbook(fname)
    if sheetindex is None:
        sheetindex = 1
    ws = wb.sheet_by_index(sheetindex)
    m_id_area = {}
    if m_idcol is None:
        m_idcol = 3
    if areacol is None:
        areacol = 0
    for row in xrange(ws.nrows):
        m_id_area[str(ws.cell(row, m_idcol).value)] = str(int(ws.cell(row, areacol).value))
    return m_id_area


def get_mediatoarea(fname=r"./xls/id.xlsx", sheetindex=1, m_idcol=3, a_idcol=0):
    bk = xlrd.open_workbook(fname)
    table = bk.sheet_by_index(sheetindex)
    m_aid = {}
    for row in xrange(table.nrows):
        if m_aid.has_key(str(int(table.cell(row,m_idcol).value))) is False:
            m_aid[str(int(table.cell(row,m_idcol).value))] = str(int(table.cell(row, a_idcol).value))
    return m_aid


def load_info(urlcol=2, soucol=0, timecol=1, fname=r'./xls/base.xls', sheetindex=0):
    '''

    :param urlcol:
    :param soucol:
    :param timecol:
    :param fname:
    :param sheetindex:
    :return:
    '''
    bk = xlrd.open_workbook(fname)
    table = bk.sheet_by_index(sheetindex)

    info = []
    for row in xrange(1, table.nrows):
        if table.cell(row, timecol).ctype == 3:
            date_value = xlrd.xldate_as_tuple(table.cell_value(row, timecol), 0)
            year = str(date_value[0])
            #time = date(*date_value[:3]).strftime('%Y/%m/%d')
        else:
            year = table.cell(row, timecol).value[:4]
        tup = (table.cell(row, soucol).value, year, str(table.cell(row, urlcol).value))
        info.append(tup)
    return info


def get_sid_year_mid_aid(info):
    '''
    :param arti_id: an int,must be given when using get_mid_sid
    :param content: a updated dict with some items already updated with m_id, s_id and m_ids
    :return:a updated dict that content[arti](also a dict) is added m_id(str) by key'm_id', s_id(list) by key's_id ',
    and m_ids(list) by key 'm_ids'
    '''
    if info is None:
        info = load_info()

    url_id = get_urltoid()
    media_id = get_mediatoid()
    urlkeys = url_id.keys()
    m_aid = get_mediatoarea()
    m_aid['unknown'] = str(10)
    mkeys = media_id.keys()
    m_id = 'no record'

    idinfo = []
    for tup in info:
        if tup[0] is not None:
            s_id = []
            try:
                for media in mkeys:
                    if media in tup[0]:
                        s_id.append(media_id[media])
            except:
                pass
        if tup[2] is '' or tup[2] is None:
            m_id = 'unknown'
        else:
            for urlkey in urlkeys:
                if str(urlkey) in tup[2]:
                    if url_id[urlkey] is None:
                        continue
                    else:
                        m_id = str(int(url_id[urlkey]))
                        break
        if m_id is not 'no record':
            a_id = m_aid[m_id]
        else:
            a_id = None
        id = (s_id, tup[1], m_id, a_id)
        idinfo.append(id)
    return idinfo


def write_result(idinfo):
    '''
    :param content: a completed dict with
    :return:
    '''
    newwb = Workbook()
    resheet = newwb.create_sheet(u"result", 0)
    for i in xrange(1, len(idinfo)+1):
        s_id = idinfo[i-1][0]
        year = idinfo[i-1][1]
        m_id = idinfo[i-1][2]
        a_id = idinfo[i-1][3]
        resheet.cell(row=i, column=1).value = ",".join(s_id)
        resheet.cell(row=i, column=2).value = year
        resheet.cell(row=i, column=3).value = m_id
        resheet.cell(row=i, column=4).value = a_id
    newwb.save(r"./xls/result.xlsx")


def arti_areacount(fname, sheetindex, timecol, areacol):
    '''

    :param fname: name of the file that contains main  original data
    :param sheetindex:the sheet index of the table containing main data
    :param timecol:the column of time
    :param m_idcol:the column if m_id
    :return:statistic data
    '''
    if fname is None:
        fname = r"./xls/result.xlsx"
    wb = xlrd.open_workbook(fname)
    if sheetindex is None:
        sheetindex = 0
    table = wb.sheet_by_index(sheetindex)
    if timecol is None:
        timecol = 1   #data = {2002:[0,0,0],
    if areacol is None:
        areacol = 3

    data = {}
    ta = []

    for row in xrange(table.nrows):
        if table.cell(row, areacol).value is '' or table.cell(row, timecol).value is '':
            continue
        tatup = (int(table.cell(row, timecol).value), int(table.cell(row, areacol).value))
        #print type(table.cell(row, timecol).value)
        ta.append(tatup)
    for oneta in xrange(len(ta)):
        if ta[oneta][0] is not None:
            year = int(ta[oneta][0])
        else:
            continue
        if data.has_key(year) is False:
            data[year] = [0, 0, 0]
    for oneta in xrange(len(ta)):

        if ta[oneta][1] == '10' or ta[oneta][1] == 10:
            data[ta[oneta][0]][0] += 1
        if ta[oneta][1] == '11' or ta[oneta][1] == 11:
            data[ta[oneta][0]][1] += 1
        if ta[oneta][1] == '12' or ta[oneta][1] == 12:
            data[ta[oneta][0]][2] += 1

    print data
    return data


def write_data():
    data = arti_areacount(None, None, None, None)
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('data')
    row0 = [u'时间', u'大陆', u'港澳台', u'境外']
    for i in xrange(len(row0)):
        ws.write(0, i, row0[i])

    column0 =data.keys()
    for i in xrange(len(column0)):
        ws.write(i+1, 0, column0[i])
        ws.write(i+1, 1, data[column0[i]][0])
        ws.write(i+1, 2, data[column0[i]][1])
        ws.write(i+1, 3, data[column0[i]][2])
    print '===='
    wb.save(r"./xls/statistic.xls")


#codes below written by sjw
def read_data(path):
        data = xlrd.open_workbook(path)
        table=data.sheets()[0]
        return table

def write_xls(filepath, indexOfsheet, start_row, art_column, content):
    print(len(content))
    xls_to_write = xlrd.open_workbook(filepath)
    excel_new  = copy(xls_to_write)
    sheet = excel_new.get_sheet(indexOfsheet)
    i = start_row
    if isinstance(content, int):
        sheet.write(i, art_column, str(content))
    else:
        for one in content:
            sheet.write(i, art_column, str(one))
            i += 1
    excel_new.save(filepath)

def get_from_xls(filepath, indexOfsheet, start_row, art_column):
     excel = xlrd.open_workbook(filepath)
     sheet = excel.sheets()[indexOfsheet]

     i = start_row
     row_content = []
     print (sheet.nrows)
     while i < sheet.nrows:
        if(len(sheet.cell(i, art_column).value) > 1):
            row_content.append(sheet.cell(i, art_column).value)
        else:
            row_content.append("9999")
        i += 1
     return row_content

def draw_line(table):
        #time_col=table.col_values(0)
        time=table.col_values(0)
        #time=[data.value for data in time_col]

        '''read data'''
        abroad_data=table.col_values(3)
        native_data=table.col_values(1)
        gangao_data=table.col_values(2)
        '''  remove title'''
        time.pop(0)
        abroad_data.pop(0)
        native_data.pop(0)
        gangao_data.pop(0)
        '''draw '''
        figure = plt.figure(1)
        plt.subplot(511)
        plt.title(u'abroad media')
        plt.plot(abroad_data)
        plt.xticks(range(len(time)),tuple(time))
        plt.subplot(513)
        plt.title(u'native media')
        plt.plot(native_data)
        plt.xticks(range(len(time)),tuple(time))
        plt.subplot(515)
        plt.title(u'HongKong or Macao area')
        plt.plot(gangao_data)
        plt.xticks(range(len(time)),tuple(time))
        plt.savefig("./images/base_count")

def draw_cluster_by_year(table):
        global color
        totle=[]
        bars=[]
        nrows = table.nrows
        ncols = table.ncols
        clusterName=table.row_values(0)
        clusterName.pop(0)
        time=table.col(0)
        time.pop(0)
        timeList=[data.value for data in time]
        for i in range(ncols-1):
                cluster_data=table.col(i+1)
                cluster_data.pop(0)
                cluster_data_List=[data.value for data in cluster_data]
                totle.append(cluster_data_List)

        new_totle=accumulation(totle)
        print len(timeList)
        print len(new_totle[0])

        while len(color)< len(timeList):
                color+=color

        for  i in range(len(new_totle)):

                bar=plt.bar(timeList, new_totle[len(new_totle)-i-1], 0.5, color=color[i], linewidth=0, align='center')
                bars.append(bar[0])

        plt.legend(tuple(bars),tuple(clusterName))


        plt.show()

def draw_cluster_totle(table):
        global color
        totle=[]
        bars=[]
        nrows = table.nrows
        ncols = table.ncols
        clusterName=[x for x in table.row_values(0)]
        data_value=[y for y in table.row_values(1)]


        while len(color)< len(clusterName):
                color+=color
        #plt.xlabel(u'class')
        #plt.ylabel(u'number')

        plt.bar(left=range(len(clusterName)),height=data_value,width=0.5,color=color[2], linewidth=0, align='center')
        plt.xticks(range(len(clusterName)),tuple(clusterName))
        plt.show()



def accumulation(totle):
        lenth=len(totle)
        new_totle=[]
        for i in range(lenth):
                if i ==0:
                        new_totle.append(totle[i])
                else:
                        new=merge(new_totle[i-1],totle[i])
                        new_totle.append(new)
        return new_totle

def merge(a,b):
        c=[a[i]+b[i] for i in range(min(len(a),len(b)))]
        return c

if __name__ == '__main__':

        idinfo = get_sid_year_mid_aid(None)
        write_result(idinfo)
        write_data()
        years = [2004, 2005, 2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015]
    #print years
        sum_each_year = [0,0,0,0,0,0,0,0,0,0,0,0]
        yearsToW = []
        time = get_from_xls("./xls/base.xls", 0, 1, 1)
        pattern = re.compile(r'\d\d\d\d')
        for one in time:
                match = pattern.search(one)
                if match:
                    year = int(match.group())
                    yearsToW.append(year)
                    i = 0
                    for oneyear in years:
                        if year == oneyear:
                            sum_each_year[i] += 1
                            break
                        i += 1

                else:
                    yearsToW.append("9999")
        
        write_xls("./xls/base.xls", 0, 1, 1, yearsToW)
        table=read_data('./xls/statistic.xls')
        draw_line(table)
